"""Excel template download + import — bulk path for offline contributors.

download_template(brand) → bytes of a multi-sheet xlsx that mirrors the form.
import_template(file_bytes) → dict[section_key, payload] ready for db.save_response.
"""
from io import BytesIO
from typing import Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import suggestions as sg

# Mapping: section_key -> (sheet name, columns, payload-extraction kind)
FIELD_COLS = ["field", "type", "options", "mandatory", "integration_needed", "data_capture_source", "conditional_rule"]
FIELD_BUILDER = {
    "partner_360":           ("Partner_360",  sg.PARTNER_360_FIELDS),
    "customer_360":          ("Customer_360", sg.CUSTOMER_360_FIELDS),
    "sales_opp_details":     ("Sales_Opp",    sg.REDHAT_OPPORTUNITY_DETAILS),
    "sales_contact_details": ("Sales_Contact",sg.REDHAT_CONTACT_DETAILS),
    "sales_deal_details":    ("Sales_Deal",   sg.REDHAT_DEAL_DETAILS),
}
APPR_COLS = ["stage", "approver", "trigger", "sla_hours", "can_revert"]
DASH_COLS = ["dashboard", "audience", "frequency"]
PEOPLE_KEYS  = ["brand_lead_name", "brand_lead_email", "roles_involved", "decision_maker",
                "daily_users", "occasional_users", "notes"]
BP_KEYS      = ["data_hygiene", "integrations_wishlist", "dedup_sops"]
NOTES_KEYS   = ["pain_points", "must_haves", "nice_to_haves", "risks", "questions_for_zoho"]


def _style_header(ws, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="262E3A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _set_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def download_template(brand: str) -> bytes:
    wb = Workbook()

    # Cover
    cov = wb.active; cov.title = "Cover"
    cov["A1"] = "REDINGTON DISCOVERY — BULK TEMPLATE"
    cov["A1"].font = Font(bold=True, size=14, color="C8102E")
    cov["A3"] = "Brand"; cov["B3"] = brand
    cov["A4"] = "Your name"; cov["B4"] = ""
    cov["A5"] = "Email"; cov["B5"] = ""
    cov["A6"] = "Role"; cov["B6"] = ""
    cov["A8"] = "Instructions"
    cov["A8"].font = Font(bold=True)
    cov["A9"]  = "1. Fill in your details above (Email is required to attribute the import)."
    cov["A10"] = "2. Fill any of the section sheets. You can leave entire sheets blank."
    cov["A11"] = "3. On field-builder sheets, mandatory/integration_needed accept Yes/No or TRUE/FALSE."
    cov["A12"] = "4. Save as .xlsx and upload via Admin → Bulk Import."
    _set_widths(cov, [22, 60])

    # Field-builder sheets — pre-filled with suggestion rows the user can edit/delete
    for sk, (sheet, defaults) in FIELD_BUILDER.items():
        ws = wb.create_sheet(sheet)
        ws.append([c.replace("_", " ").title() for c in FIELD_COLS])
        _style_header(ws, len(FIELD_COLS))
        for row in defaults:
            ws.append([row.get(c, "") for c in FIELD_COLS])
        _set_widths(ws, [32, 14, 38, 12, 18, 22, 38])

    # Approvals
    ap = wb.create_sheet("Approvals")
    ap.append([c.replace("_", " ").title() for c in APPR_COLS])
    _style_header(ap, len(APPR_COLS))
    for s in sg.APPROVAL_STAGE_SUGGESTIONS:
        ap.append([s.get(c, "") for c in APPR_COLS])
    _set_widths(ap, [28, 24, 50, 12, 12])

    # Dashboards
    da = wb.create_sheet("Dashboards")
    da.append([c.replace("_", " ").title() for c in DASH_COLS])
    _style_header(da, len(DASH_COLS))
    for d in sg.DASHBOARD_SUGGESTIONS:
        da.append([d.get(c, "") for c in DASH_COLS])
    _set_widths(da, [40, 40, 18])

    # People (key/value)
    pe = wb.create_sheet("People")
    pe.append(["Key", "Value"]); _style_header(pe, 2)
    for k in PEOPLE_KEYS: pe.append([k, ""])
    _set_widths(pe, [24, 60])

    # Best Practices + Open Notes (key/value)
    bp = wb.create_sheet("Best_Practices")
    bp.append(["Key", "Value"]); _style_header(bp, 2)
    for k in BP_KEYS: bp.append([k, ""])
    _set_widths(bp, [28, 80])

    on = wb.create_sheet("Open_Notes")
    on.append(["Key", "Value"]); _style_header(on, 2)
    for k in NOTES_KEYS: on.append([k, ""])
    _set_widths(on, [28, 80])

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


# ---------- Import ----------

def _bool(v) -> bool:
    if isinstance(v, bool): return v
    s = str(v or "").strip().lower()
    return s in ("yes", "y", "true", "1", "t")

def _to_int(v) -> int | None:
    try: return int(v) if v not in (None, "") else None
    except Exception: return None


def import_template(file_bytes: bytes) -> dict[str, Any]:
    """Parse an uploaded workbook → {sections: {key: payload}, contributor: {name, email, role}}."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    out: dict[str, Any] = {"contributor": {}, "sections": {}}

    # Contributor info from Cover sheet
    if "Cover" in wb.sheetnames:
        cov = wb["Cover"]
        out["contributor"] = {
            "name":  str(cov["B4"].value or "").strip(),
            "email": str(cov["B5"].value or "").strip(),
            "role":  str(cov["B6"].value or "").strip() or "Other",
        }

    # Field-builder sheets
    for sk, (sheet, _) in FIELD_BUILDER.items():
        if sheet in wb.sheetnames:
            fields = _read_field_sheet(wb[sheet])
            if fields:
                out["sections"][sk] = {"fields": fields, "notes": ""}

    # Approvals
    if "Approvals" in wb.sheetnames:
        stages = _read_approvals_sheet(wb["Approvals"])
        if stages:
            out["sections"]["approvals"] = {"stages": stages, "escalation": "", "notes": ""}

    # Dashboards
    if "Dashboards" in wb.sheetnames:
        dashes = _read_dashboards_sheet(wb["Dashboards"])
        if dashes:
            out["sections"]["dashboards"] = {"dashboards": dashes, "notes": ""}

    # Key/value sheets
    for sheet_name, sk in [("People", "people"), ("Best_Practices", "best_practices"), ("Open_Notes", "open_notes")]:
        if sheet_name in wb.sheetnames:
            kv = _read_kv_sheet(wb[sheet_name])
            if any(v for v in kv.values()):
                if sheet_name == "People" and "roles_involved" in kv and isinstance(kv["roles_involved"], str):
                    kv["roles_involved"] = [r.strip() for r in kv["roles_involved"].split(",") if r.strip()]
                out["sections"][sk] = kv

    return out


def _read_field_sheet(ws) -> list[dict]:
    out: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or all(c in (None, "") for c in row): continue
        d = {
            "field":              str(row[0] or "").strip() if len(row) > 0 else "",
            "type":               str(row[1] or "").strip() if len(row) > 1 else "",
            "options":            str(row[2] or "").strip() if len(row) > 2 else "",
            "mandatory":          _bool(row[3]) if len(row) > 3 else False,
            "integration_needed": _bool(row[4]) if len(row) > 4 else False,
            "data_capture_source":str(row[5] or "").strip() if len(row) > 5 else "",
            "conditional_rule":   str(row[6] or "").strip() if len(row) > 6 else "",
        }
        if d["field"]:
            out.append(d)
    return out


def _read_approvals_sheet(ws) -> list[dict]:
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or all(c in (None, "") for c in row): continue
        d = {
            "stage":      str(row[0] or "").strip() if len(row) > 0 else "",
            "approver":   str(row[1] or "").strip() if len(row) > 1 else "",
            "trigger":    str(row[2] or "").strip() if len(row) > 2 else "",
            "sla_hours":  _to_int(row[3]) if len(row) > 3 else None,
            "can_revert": _bool(row[4]) if len(row) > 4 else False,
        }
        if d["stage"]: out.append(d)
    return out


def _read_dashboards_sheet(ws) -> list[dict]:
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or all(c in (None, "") for c in row): continue
        d = {
            "dashboard": str(row[0] or "").strip() if len(row) > 0 else "",
            "audience":  str(row[1] or "").strip() if len(row) > 1 else "",
            "frequency": str(row[2] or "").strip() if len(row) > 2 else "",
        }
        if d["dashboard"]: out.append(d)
    return out


def _read_kv_sheet(ws) -> dict:
    out: dict = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row or not row[0]: continue
        k = str(row[0]).strip()
        v = row[1] if len(row) > 1 else None
        out[k] = "" if v is None else str(v).strip()
    return out
